import requests
from bs4 import BeautifulSoup as bs
from collections import namedtuple
import csv
from openpyxl import load_workbook
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from decouple import config

options = webdriver.ChromeOptions()
# options.add_argument('--headless')
driver = webdriver.Chrome(executable_path="C:\\Users\\Bishal\\Downloads\chromedriver_win32\\chromedriver.exe", options=options)

class Linkedin:
  mainUrl = 'https://www.linkedin.com'
  LINKEDIN_EMAIL = config('LINKEDIN_EMAIL', '')
  LINKEDIN_PASSWORD = config('LINKEDIN_PASSWORD', '')

  email = LINKEDIN_EMAIL
  password = LINKEDIN_PASSWORD
  def url(title, location):
    return 'https://www.linkedin.com/jobs/search?keywords=' + title + '&location=' + location + '&geoId=&trk=homepage-jobseeker_jobs-search-bar_search-submit&position=1&pageNum=0'
  driver.get("https://www.linkedin.com")
  driver.find_element_by_id('session_key').send_keys(email)
  el = driver.find_element_by_id('session_password')
  el.send_keys(password)
  el.send_keys(Keys.ENTER)
  # Scrape('developer', 'dubai', 'output.xlsx', 'acs')

  def Scrape(title, loc, output, date_posted):
    driver.get(Linkedin.url(title, loc))
    print('Scrapping this url ' + Linkedin.url(title, loc))
    job_cards = bs(driver.page_source, features="lxml").find_all('li', {'class': 'jobs-search-results__list-item'})
    for job in job_cards:
      try:
        job_title = job.find('a', {'class': 'job-card-list__title'}).text.strip()
      except:
        job_title = ''
      try:
        company = job.find('a', {'class': 'job-card-container__company-name'}).text.strip()
      except:
        company = ''
      try:
        location = job.find('li', {'class': 'job-card-container__metadata-item'}).text.strip()
      except:
        location = ''
      try:
        job_summary = ''
      except:
        job_summary = ''
      try:
        post_date = job.find('time').get_text()
      except:
        post_date = ''
      try:
        job_url = Linkedin.mainUrl  + job.find('a', {'class': 'job-card-list__title'})['href']
      except:
        job_url = ''

      record = job_title, company, location, job_summary, post_date, job_url
      Linkedin.save_record_to_csv(record, output)

  def save_record_to_csv(record, filepath, create_new_file=False):
    header = ["JobTitle", "Company", "Location", "Summary", "PostDate", "JobUrl"]
    if create_new_file:
      wb = load_workbook(filename=filepath)
      wb.remove(wb.worksheets[0])
      wb.create_sheet()
      ws = wb.worksheets[0]
      ws.append(header)
      wb.save(filepath)
    else:
      wb = load_workbook(filename=filepath)
      # Select First Worksheet
      ws = wb.worksheets[0]
      ws.append(record)
      wb.save(filepath)


# Linkedin.Scrape('developer', 'dubai', 'output.xlsx','acs')


