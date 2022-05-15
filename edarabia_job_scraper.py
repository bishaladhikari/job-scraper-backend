from bs4 import BeautifulSoup as bs
import requests
from random import random
from time import sleep
from email.message import EmailMessage
from collections import namedtuple
import smtplib
import csv
from openpyxl import load_workbook
import os
from datetime import datetime


def save_record_to_csv(record, filepath, create_new_file=False):
    """Save an individual record to file; set `new_file` flag to `True` to generate new file"""
    header = ["JobTitle", "Company", "Location", "Summary", "PostDate", "JobUrl"]
    if create_new_file:
        wb = load_workbook(filename=filepath)
        wb.remove(wb.worksheets[0])
        wb.create_sheet()
        ws = wb.worksheets[0]
        ws.append(header)
        wb.save(filepath)
    else:
        wb = load_workbook(filename=filepath)
        # Select First Worksheet
        ws = wb.worksheets[0]
        ws.append(record)
        wb.save(filepath)

def getMoreInfo(url):
    try:
        newresp = requests.get(url, headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT x.y; rv:10.0) Gecko/20100101 Firefox/10.0',
        })
        moreInfoSoup = bs(newresp.content, features="lxml")
        ulList = moreInfoSoup.find('ul', {'class' : 'custom-fields'})
        company = ulList.findAll("li")[0].find('a')['title']
        address = moreInfoSoup.find("ul", {'class' : 'list-items-top'}).text.replace('Address: ', '')
        return company, address
    except:
        a = "__"
        b = "__"
        return a, b 

def scrapeJob(url, title):
    # HERE THIS FUNCTION SCRAPES ALL JOBS AND STORES IN A LIST CALLED JOBS
    resp = requests.post(url, data={
        'postType': 'job',
        'keyword': title,
        'country_city': ''
    })
    soup = bs(resp.content, features="lxml")
    jobCards = soup.find_all('div', "list-items")
    jobs = []
    for jobCard in jobCards:
        title = jobCard.findAll('a')[0]['title']
        link = jobCard.findAll('a')[0]['href']
        country = jobCard.findAll('a')[2]['title']
        company, address = getMoreInfo(link)
        jobs.append({'title' : title, 'link' : link , 'location' : address, 'company' : company})
    return jobs


def main(domain, date_posted, job_title, job_location, filepath, email=None):
    job_base_url = 'https://www.edarabia.com/wp-content/themes/newed/contents/custom-search.php'
    # HERE WE ARE ASKING FOR JOBS TO OUR BROTHER,
    # AND HE RETURNS US A LIST AS JOBS,
    jobs = scrapeJob(job_base_url, job_title)
    # NOW WE ARE ADDING THAT DATA TO OUR EXCEL SHEET AND RETURNING,
    for job in jobs:
        result = job['title'], job['company'], job['location'], '', 'few days ago', job['link']
        save_record_to_csv(result, filepath)


